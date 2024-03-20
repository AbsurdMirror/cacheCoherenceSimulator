
class DataDriver:
    drive_data = 0
    verify_data = 0
    verify_data_valid = False
    print_driver = None

    verify_times = 0

    cores = []

    last_verify_tick = 0

    def get_data(self):
        self.drive_data += 1
        return self.drive_data

    def data_init(self, data_value):
        self.drive_data = data_value
        self.verify_data = data_value
        self.verify_data_valid = True

    def connect(self, objects):
        for core in objects:
            core.data_driver = self
            self.cores.append(core)

    def verify_load(self, core):
        if  (not self.verify_data_valid) or core.load_data != self.verify_data:
            self.print_driver.print("ERROR: Data mismatch", Flag="Verify")
            raise Exception("Should not happen")
            exit()
        else:
            self.print_driver.print(f"verify: Data match: {core.name()} load {core.load_data}", Flag="Verify")
        core.verity_load_times += 1
        self.last_verify_tick = self.tick_driver.tick
        self.verify_times += 1
    
    def verify_store(self, core):
        self.verify_data = core.store_data
        self.verify_data_valid = True
        self.print_driver.print(f"verify: Data update: {core.name()} store {core.store_data}", Flag="Verify")
        core.verity_store_times += 1
        self.last_verify_tick = self.tick_driver.tick
        self.verify_times += 1

    def tick_run(self):
        if self.tick_driver.tick - self.last_verify_tick > 100:
            print(f"WARNING: verify not happen more than {100} Ticks")


class TickDriver:
    tick = 0
    max_tick = 1024
    print_driver = None

    tick_objects = []

    def add_objects(self, objects):
        for object in objects:
            self.tick_objects.append(object)
            object.tick_driver = self

    def name(self):
        return "Tick"

    def run(self):
        while self.tick < self.max_tick:
            for object in self.tick_objects:
                object.tick_run()
            self.tick += 1
            # print("")
        print(f"simulation finished @ {self.tick} Ticks")
        self.excel_driver.ptrace(self, f"{self.tick}")

class PrintDriver:

    tick_driver = None

    print_flags = []

    def add_flags(self, flags):
        for flag in flags:
            self.print_flags.append(flag)

    def add_tick_driver(self, tick_driver):
        self.tick_driver = tick_driver
    
    def connect(self, objects):
        for object in objects:
            object.print_driver = self
    
    def print(self, *args, **kwargs):  
        # 如果有flag参数但是不在flags数组中，则不打印
        flag = kwargs.pop('Flag', None)
        if not (flag is None or flag in self.print_flags):  
            return

        prefix = str(self.tick_driver.tick) + " Tick: "
        # 使用format方法将前缀和参数格式化输出  
        formatted_args = [prefix] + list(args)  
        formatted_str = " ".join(map(str, formatted_args))  
          
        # 分离出end参数，因为print的end参数在格式化字符串时不需要  
        end = kwargs.pop('end', '\n')  
          
        # 使用内置的print函数打印格式化后的字符串，并传递剩余的kwargs  
        print(formatted_str, end=end, **kwargs)  
    
    def enum_str(self, enum_value):
        return f"{enum_value.__class__.__name__}.{enum_value.name}"

    def print_snh(self, node_name, event, state):  
        self.print(f"ERROR: should not happen: From ({self.enum_str(event.cmd)}@{event.requestor.name()}) To ({self.enum_str(state)}@{node_name})")

import openpyxl  

class ExcelManager:  
    def __init__(self, path):  
        self.path = path  
        self.workbook = openpyxl.Workbook()  

    def save_workbook(self):  
        self.workbook.save(self.path)  
  
    def create_sheet(self, sheet_name):  
        self.workbook.create_sheet(sheet_name)  
  
    def get_sheet(self, sheet_name):  
        return self.workbook[sheet_name]  
  
    def read_cell(self, sheet_name, row, column):  
        sheet = self.get_sheet(sheet_name)  
        cell_value = sheet.cell(row=row, column=column).value  
        return cell_value  
  
    def write_cell(self, sheet_name, row, column, value):  
        sheet = self.get_sheet(sheet_name)  
        sheet.cell(row=row, column=column).value = value  

class ExcelDriver(ExcelManager):
    def __init__(self, path, cache_states, cache_event_cmds, llc_states, llc_event_cmds):
        super().__init__(path)
        self.create_sheet("trace")
        self.create_sheet("cache_state")
        self.create_sheet("llc_state")

        self.cache_state_num = len(cache_states)
        self.cache_event_cmd_num = len(cache_event_cmds)
        self.llc_state_num = len(llc_states)
        self.llc_event_cmd_num = len(llc_event_cmds)

        for event_cmd in cache_event_cmds:
            self.write_cell("cache_state", 1, event_cmd.value + 2, event_cmd.name)
        for state in cache_states:
            self.write_cell("cache_state", state.value + 2, 1, state.name)

        for event_cmd in llc_event_cmds:
            self.write_cell("llc_state", 1, event_cmd.value + 2, event_cmd.name)
        for state in llc_states:
            self.write_cell("llc_state", state.value + 2, 1, state.name)

        cache_state_table = [[None] * self.cache_event_cmd_num for _ in range(self.cache_state_num)]  
        llc_state_table = [[None] * self.llc_event_cmd_num for _ in range(self.llc_state_num)]
        self.state_table = {
            "cache": cache_state_table,
            "llc": llc_state_table
        }
        super().save_workbook()

    def ptrace(self, node, value):
        self.write_cell("trace", self.tick_driver.tick + 2, node.column, value)

    def bind_nodes(self, nodes):
        column = 1
        for node in nodes:
            node.column = column
            column += 1
            self.write_cell("trace", 1, node.column, node.name())
            node.excel_driver = self

    def bind_tick_driver(self, tick_driver):
        self.tick_driver = tick_driver

    def pstate(self, node_type, state, event_cmd, tick):
        if self.state_table[node_type][state][event_cmd] is None:  
            self.state_table[node_type][state][event_cmd] = tick
    
    def save_workbook(self): 
        for state in range(self.cache_state_num):  
            for event_cmd in range(self.cache_event_cmd_num):  
                cell_value = self.state_table["cache"][state][event_cmd]  
                if cell_value is None:  
                    cell_value = "SNH"
                self.write_cell("cache_state", state + 2, event_cmd + 2, cell_value)

        for state in range(self.llc_state_num):  
            for event_cmd in range(self.llc_event_cmd_num):  
                cell_value = self.state_table["llc"][state][event_cmd]  
                if cell_value is None:  
                    cell_value = "SNH"
                self.write_cell("llc_state", state + 2, event_cmd + 2, cell_value)

        super().save_workbook()
